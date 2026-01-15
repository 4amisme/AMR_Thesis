from pathlib import Path
import pandas as pd

# ใช้ openpyxl สำหรับอ่าน/เขียน Excel
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows

# CONFIG 
xlsx_path = Path("references/Org_Map-15-Jan.xlsx")      # <-- ใส่ชื่อไฟล์ของคุณ
source_sheet = "Organism"                # <-- ชื่อชีตที่มีตารางต้นฉบับ
new_sheet_name = "Organism_Mapping"    # <-- ชื่อชีตใหม่ที่จะบันทึก

# อ่านข้อมูลจากชีตต้นทาง
df = pd.read_excel(xlsx_path, sheet_name=source_sheet, dtype=str)

# คาดหวังคอลัมน์ชื่อ "Organism" และ "Code_Org"
required_cols = {"Organism", "Code_Org"}
missing = required_cols - set(df.columns)
if missing:
    raise ValueError(f"Missing required columns: {missing}. Found columns: {list(df.columns)}")

# ทำความสะอาด + แตก Code_Org เป็นหลายแถว (long format)
df2 = df.copy()
df2["Organism"] = df2["Organism"].astype(str).str.strip()
df2["Code_Org"] = df2["Code_Org"].astype(str).fillna("").str.strip()

# แยกด้วย comma แล้ว trim ช่องว่าง
df2["Alias_List"] = df2["Code_Org"].apply(
    lambda s: [x.strip() for x in s.split(",") if x.strip()]
)

mapping_long = (
    df2[["Organism", "Alias_List"]]
    .explode("Alias_List", ignore_index=True)
    .rename(columns={"Alias_List": "Alias_Code"})
)

mapping_long = mapping_long[mapping_long["Alias_Code"].notna() & (mapping_long["Alias_Code"] != "")]

# เขียนกลับเข้าไฟล์เดิมเป็น worksheet ใหม่
wb = load_workbook(xlsx_path)

# ถ้าชีตชื่อซ้ำอยู่แล้ว ให้ลบทิ้งก่อน (เพื่อรันซ้ำได้)
if new_sheet_name in wb.sheetnames:
    ws_old = wb[new_sheet_name]
    wb.remove(ws_old)

ws = wb.create_sheet(title=new_sheet_name)

# ใส่ header + data
for r in dataframe_to_rows(mapping_long, index=False, header=True):
    ws.append(r)

# ปรับความกว้างคอลัมน์ให้อ่านง่าย
ws.column_dimensions["A"].width = 35  # Organism
ws.column_dimensions["B"].width = 18  # Alias_Code

wb.save(xlsx_path)

print(f"Done! Wrote {len(mapping_long):,} rows to sheet '{new_sheet_name}' in: {xlsx_path}")
